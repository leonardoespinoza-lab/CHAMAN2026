#!/usr/bin/env python3
import argparse
import csv
import json
import os
import re
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ARGENTINA_BOUNDS = {
    "lat_min": -56.0,
    "lat_max": -21.0,
    "lon_min": -74.0,
    "lon_max": -53.0,
}

USER_AGENT = (
    "ChamanAgro/1.0 Corteva-distributors-geocoding "
    "(operational audit; contact: operaciones@chaman.ag)"
)

GOOGLE_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY") or os.environ.get("MAPS_KEY") or ""


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize(value):
    text = clean(value).upper()
    text = "".join(
        ch
        for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )
    text = re.sub(r"[^A-Z0-9]+", " ", text).strip()
    text = re.sub(r"\s+", " ", text)
    aliases = {
        "WENSESLAO": "WENCESLAO",
        "SALLIQUELLO": "SALLIQUELO",
        "CAPITAL FEDERAL": "CIUDAD AUTONOMA BUENOS AIRES",
        "CABA": "CIUDAD AUTONOMA BUENOS AIRES",
        "GRAL ": "GENERAL ",
        "GRAL.": "GENERAL",
        "CNEL ": "CORONEL ",
        "CNEL.": "CORONEL",
    }
    for old, new in aliases.items():
        text = text.replace(old, new)
    return re.sub(r"\s+", " ", text).strip()


LOCALITY_DISPLAY_ALIASES = {
    ("ALDERETE TUCUMAN", "TUCUMAN"): "Alderetes",
    ("SANTA ROSA DE RIO PRIMERO", "CORDOBA"): "Villa Santa Rosa",
    ("GRAL PICO", "LA PAMPA"): "General Pico",
    ("GENERAL PICO", "LA PAMPA"): "General Pico",
    ("CASTELLI", "CHACO"): "Juan Jose Castelli",
    ("CASA CENTRAL", "CORDOBA"): "Villa Maria",
    ("GRAL BELGRANO", "BUENOS AIRES"): "General Belgrano",
    ("GENERAL BELGRANO", "BUENOS AIRES"): "General Belgrano",
}


def canonical_locality(localidad, provincia):
    loc_norm = normalize(localidad)
    prov_norm = normalize(provincia)
    return LOCALITY_DISPLAY_ALIASES.get((loc_norm, prov_norm), clean(localidad))


def title_case(value):
    text = clean(value)
    words = []
    keep_upper = {"CP", "RN", "RP", "S/N", "SRL", "S.A.", "SA"}
    for word in text.split(" "):
        if not word:
            continue
        if word.upper() in keep_upper:
            words.append(word.upper())
        else:
            words.append(word[:1].upper() + word[1:].lower())
    return " ".join(words)


def to_float(value):
    try:
        if value is None or clean(value) == "":
            return None
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def decode_url_text(value):
    text = clean(value)
    if not text:
        return ""
    previous = text
    for _ in range(2):
        decoded = urllib.parse.unquote_plus(previous)
        if decoded == previous:
            break
        previous = decoded
    return previous


def extract_coordinates_from_text(value):
    text = decode_url_text(value)
    if not text:
        return None
    patterns = [
        r"@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)",
        r"[?&](?:q|query|ll|destination)=(-?\d+(?:\.\d+)?),\s*(-?\d+(?:\.\d+)?)",
        r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)",
        r"/dir/[^/]+/(-?\d+(?:\.\d+)?),\s*(-?\d+(?:\.\d+)?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        lat = to_float(match.group(1))
        lon = to_float(match.group(2))
        if in_argentina(lat, lon):
            return lat, lon
    return None


def extract_query_from_google_url(value):
    text = clean(value)
    if not text:
        return ""
    decoded = decode_url_text(text)
    if not decoded.startswith("http"):
        return decoded
    try:
        parsed = urllib.parse.urlparse(text)
        query = urllib.parse.parse_qs(parsed.query)
        for key in ["query", "q", "destination"]:
            if query.get(key):
                return clean(query[key][0])
    except Exception:
        return ""
    return ""


def in_argentina(lat, lon):
    return (
        lat is not None
        and lon is not None
        and ARGENTINA_BOUNDS["lat_min"] <= lat <= ARGENTINA_BOUNDS["lat_max"]
        and ARGENTINA_BOUNDS["lon_min"] <= lon <= ARGENTINA_BOUNDS["lon_max"]
    )


def infer_location(localidad, provincia):
    localidad_limpia = clean(localidad)
    provincia_limpia = clean(provincia)
    loc_norm = normalize(localidad_limpia)
    if not provincia_limpia:
        province_aliases = {
            "BUENOS AIRES": "BUENOS AIRES",
            "CHACO": "CHACO",
            "TUCUMAN": "TUCUMAN",
            "CORDOBA": "CORDOBA",
            "SANTA FE": "SANTA FE",
            "SALTA": "SALTA",
            "SANTIAGO DEL ESTERO": "SANTIAGO DEL ESTERO",
        }
        for token, inferred in province_aliases.items():
            if token in loc_norm:
                provincia_limpia = inferred
                loc_norm = loc_norm.replace(token, " ").strip()
                localidad_limpia = title_case(loc_norm)
                break
    localidad_limpia = re.sub(r",?\s*\d{3,5}$", "", localidad_limpia).strip()
    localidad_limpia = canonical_locality(localidad_limpia, provincia_limpia)
    return localidad_limpia, provincia_limpia


def build_address(row):
    domicilio = clean(row.get("DOMICILIO"))
    localidad, provincia = infer_location(row.get("LOCALIDAD"), row.get("PROVINCIA"))
    parts = [domicilio, localidad, provincia, "Argentina"]
    return ", ".join([p for p in parts if p])


def build_base_names(records):
    names = []
    for row in records:
        distribuidor = clean(row.get("DISTRIBUIDOR"))
        localidad, _ = infer_location(row.get("LOCALIDAD"), row.get("PROVINCIA"))
        localidad = title_case(localidad)
        names.append(f"{distribuidor} - {localidad}" if localidad else distribuidor)
    return names


def raw_match_key(distribuidor, domicilio, localidad, provincia):
    return "|".join(
        [
            normalize(distribuidor),
            normalize(domicilio),
            normalize(localidad),
            normalize(provincia),
        ]
    )


def load_mymaps_points(path):
    if not path:
        return {}
    points_path = Path(path)
    if not points_path.exists():
        raise FileNotFoundError(f"No existe archivo My Maps: {points_path}")
    points = json.loads(points_path.read_text(encoding="utf-8"))
    index = {}
    for point in points:
        lat = to_float(point.get("lat"))
        lon = to_float(point.get("lon"))
        if not in_argentina(lat, lon):
            continue
        key = raw_match_key(
            point.get("distribuidor"),
            point.get("domicilio"),
            point.get("localidad"),
            point.get("provincia"),
        )
        if key:
            index[key] = {**point, "lat": lat, "lon": lon}
    return index


def make_unique_names(records):
    names = build_base_names(records)
    counts = Counter(normalize(name) for name in names)
    with_province = []
    for row, name in zip(records, names):
        if counts[normalize(name)] > 1:
            _, provincia = infer_location(row.get("LOCALIDAD"), row.get("PROVINCIA"))
            if provincia:
                name = f"{name} - {title_case(provincia)}"
        with_province.append(name)

    counts2 = Counter(normalize(name) for name in with_province)
    final = []
    for row, name in zip(records, with_province):
        if counts2[normalize(name)] > 1:
            domicilio = title_case(row.get("DOMICILIO"))[:55].strip()
            if domicilio:
                name = f"{name} - {domicilio}"
        final.append(name)

    counts3 = Counter(normalize(name) for name in final)
    for idx, name in enumerate(final):
        if counts3[normalize(name)] > 1:
            final[idx] = f"{name} - Fila {records[idx]['_excel_row']}"
    return final


def load_records(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Distribuidores con Maps"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [clean(cell) if cell is not None else f"COL_{i + 1}" for i, cell in enumerate(rows[0])]
    records = []
    for excel_row, row in enumerate(rows[1:], start=2):
        if not any(cell is not None and clean(cell) for cell in row):
            continue
        record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        record["_excel_row"] = excel_row
        records.append(record)
    return records


def cache_key(provider, query):
    text = clean(query).lower()
    text = "".join(
        ch
        for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )
    text = re.sub(r"\s+", " ", text).strip()
    return f"{provider}:{text}"


def legacy_cache_key(provider, query):
    return f"{provider}:{normalize(query)}"


def can_use_legacy_cache(query):
    raw = clean(query).upper()
    # These corrected aliases intentionally must not reuse failed cache entries
    # from the misspelled original Excel text.
    if "WENCESLAO" in raw or "SALLIQUELO" in raw or "CIUDAD AUTONOMA" in normalize(raw):
        return False
    return True


def http_json(url, timeout=25, attempts=3):
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    last_error = None
    for attempt in range(attempts):
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            last_error = error
            if error.code in {429, 500, 502, 503, 504} and attempt < attempts - 1:
                time.sleep(4 + attempt * 6)
                continue
            raise
        except Exception as error:
            last_error = error
            if attempt < attempts - 1:
                time.sleep(3 + attempt * 4)
                continue
            raise
    raise last_error


def http_post_json(url, body, headers=None, timeout=25, attempts=3):
    data = json.dumps(body, ensure_ascii=False).encode("utf-8")
    request_headers = {
        "User-Agent": USER_AGENT,
        "Content-Type": "application/json",
    }
    if headers:
        request_headers.update(headers)
    request = urllib.request.Request(url, data=data, headers=request_headers, method="POST")
    last_error = None
    for attempt in range(attempts):
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            last_error = error
            if error.code in {429, 500, 502, 503, 504} and attempt < attempts - 1:
                time.sleep(4 + attempt * 6)
                continue
            raise
        except Exception as error:
            last_error = error
            if attempt < attempts - 1:
                time.sleep(3 + attempt * 4)
                continue
            raise
    raise last_error


def nominatim_search(query, cache, delay):
    key = cache_key("nominatim", query)
    if key in cache:
        return cache[key]
    old_key = legacy_cache_key("nominatim", query)
    if can_use_legacy_cache(query) and old_key in cache:
        cache[key] = cache[old_key]
        return cache[key]
    url = (
        "https://nominatim.openstreetmap.org/search?"
        + urllib.parse.urlencode(
            {
                "format": "jsonv2",
                "limit": 5,
                "countrycodes": "ar",
                "addressdetails": 1,
                "q": query,
            }
        )
    )
    try:
        result = http_json(url)
    except Exception as error:
        result = {"error": repr(error)}
    cache[key] = result
    time.sleep(delay)
    return result


def google_cache_key(provider, query):
    return cache_key(provider, query)


def google_address_component(result, component_type):
    for component in result.get("address_components") or []:
        if component_type in component.get("types", []):
            return component.get("long_name") or ""
    return ""


def google_result_matches(result, localidad, provincia):
    formatted = normalize(result.get("formatted_address") or result.get("formattedAddress") or "")
    provincia_norm = normalize(provincia)
    localidad_norm = normalize(localidad)
    state = normalize(google_address_component(result, "administrative_area_level_1"))
    locality_values = [
        google_address_component(result, "locality"),
        google_address_component(result, "postal_town"),
        google_address_component(result, "administrative_area_level_2"),
        google_address_component(result, "sublocality"),
    ]
    locality_text = normalize(" ".join(locality_values))

    if provincia_norm:
        if state and provincia_norm not in state:
            if not (provincia_norm == "BUENOS AIRES" and "CIUDAD AUTONOMA BUENOS AIRES" in state):
                return False
        elif provincia_norm not in formatted and not (
            provincia_norm == "BUENOS AIRES" and "CIUDAD AUTONOMA BUENOS AIRES" in formatted
        ):
            return False

    if localidad_norm and localidad_norm not in {"CASA CENTRAL"}:
        locality_tokens = [token for token in localidad_norm.split(" ") if len(token) > 2]
        compare_text = f"{locality_text} {formatted}"
        if locality_tokens and not any(token in compare_text for token in locality_tokens[:3]):
            return False
    return True


def google_places_search(query, localidad, provincia, cache, delay):
    if not GOOGLE_API_KEY or not query:
        return None
    key = google_cache_key("google_places", query)
    if key in cache:
        result = cache[key]
    else:
        url = "https://places.googleapis.com/v1/places:searchText"
        body = {
            "textQuery": query,
            "regionCode": "AR",
            "maxResultCount": 3,
        }
        try:
            result = http_post_json(
                url,
                body,
                headers={
                    "X-Goog-Api-Key": GOOGLE_API_KEY,
                    "X-Goog-FieldMask": (
                        "places.formattedAddress,places.displayName,"
                        "places.location,places.addressComponents"
                    ),
                },
            )
        except Exception as error:
            result = {"error": repr(error)}
        cache[key] = result
        time.sleep(delay)

    places = (result.get("places") or []) if isinstance(result, dict) else []
    for place in places:
        location = place.get("location") or {}
        lat = to_float(location.get("latitude"))
        lon = to_float(location.get("longitude"))
        if not in_argentina(lat, lon):
            continue
        adapted = {
            "formatted_address": place.get("formattedAddress") or "",
            "address_components": [
                {
                    "long_name": item.get("longText") or item.get("shortText") or "",
                    "types": item.get("types") or [],
                }
                for item in place.get("addressComponents") or []
            ],
        }
        if not google_result_matches(adapted, localidad, provincia):
            continue
        name = (place.get("displayName") or {}).get("text") or ""
        formatted = place.get("formattedAddress") or ""
        return {
            "lat": lat,
            "lon": lon,
            "estado": "GEOCODIFICADA_GOOGLE_PLACES",
            "confianza": "alta",
            "fuente": "Google Places Text Search",
            "consulta": query,
            "resultado": clean(" - ".join([name, formatted]).strip(" -")),
        }
    return None


def google_geocode(query, localidad, provincia, cache, delay):
    if not GOOGLE_API_KEY or not query:
        return None
    key = google_cache_key("google_geocode", query)
    if key in cache:
        result = cache[key]
    else:
        url = (
            "https://maps.googleapis.com/maps/api/geocode/json?"
            + urllib.parse.urlencode(
                {
                    "address": query,
                    "region": "ar",
                    "components": "country:AR",
                    "key": GOOGLE_API_KEY,
                }
            )
        )
        try:
            result = http_json(url)
        except Exception as error:
            result = {"error": repr(error)}
        cache[key] = result
        time.sleep(delay)

    if not isinstance(result, dict) or result.get("status") not in {"OK", "ZERO_RESULTS"}:
        return None
    for item in result.get("results") or []:
        location = ((item.get("geometry") or {}).get("location") or {})
        lat = to_float(location.get("lat"))
        lon = to_float(location.get("lng"))
        if not in_argentina(lat, lon):
            continue
        if not google_result_matches(item, localidad, provincia):
            continue
        partial = item.get("partial_match") is True
        return {
            "lat": lat,
            "lon": lon,
            "estado": "GEOCODIFICADA_GOOGLE",
            "confianza": "media" if partial else "alta",
            "fuente": "Google Geocoding API",
            "consulta": query,
            "resultado": clean(item.get("formatted_address")),
        }
    return None


def province_matches(result, provincia):
    if not provincia:
        return True
    address = result.get("address") or {}
    state = normalize(address.get("state") or "")
    display = normalize(result.get("display_name"))
    provincia_norm = normalize(provincia)
    if state:
        if provincia_norm in state:
            return True
        if provincia_norm == "BUENOS AIRES" and "CIUDAD AUTONOMA BUENOS AIRES" in state:
            return True
        return False
    if provincia_norm in display:
        return True
    if provincia_norm == "BUENOS AIRES" and "CIUDAD AUTONOMA BUENOS AIRES" in (state + " " + display):
        return True
    return False


def locality_matches(result, localidad):
    if not localidad:
        return True
    localidad_norm = normalize(localidad)
    if localidad_norm in {"CASA CENTRAL"}:
        return True
    display = normalize(result.get("display_name"))
    if localidad_norm in {"CIUDAD AUTONOMA BUENOS AIRES"}:
        address = result.get("address") or {}
        state = normalize(address.get("state") or "")
        city = normalize(address.get("city") or address.get("town") or "")
        return (
            "CIUDAD AUTONOMA BUENOS AIRES" in state
            or "CIUDAD AUTONOMA BUENOS AIRES" in display
            or (city == "BUENOS AIRES" and "COMUNA" in display)
        )
    locality_tokens = [token for token in localidad_norm.split(" ") if len(token) > 2]
    if not locality_tokens:
        return True
    return any(token in display for token in locality_tokens[:3])


def is_locality_result(result):
    category = clean(result.get("category"))
    result_type = clean(result.get("type"))
    addresstype = clean(result.get("addresstype"))
    if category == "boundary" and result_type == "administrative":
        return True
    return addresstype in {
        "city",
        "town",
        "village",
        "municipality",
        "city_district",
        "state_district",
        "locality",
    }


def select_nominatim_result(results, localidad, provincia, mode="address"):
    if not isinstance(results, list):
        return None
    for result in results:
        lat = to_float(result.get("lat"))
        lon = to_float(result.get("lon"))
        if not in_argentina(lat, lon):
            continue
        if not province_matches(result, provincia):
            continue
        if not locality_matches(result, localidad):
            continue
        if mode == "locality" and not is_locality_result(result):
            continue
        return result
    return None


def classify_nominatim_result(result, mode):
    category = clean(result.get("category"))
    addresstype = clean(result.get("addresstype"))
    place_rank = int(to_float(result.get("place_rank")) or 0)
    result_type = clean(result.get("type"))
    if mode == "locality":
        return "CENTROIDE_LOCALIDAD", "media"
    if category == "highway" or addresstype in {"road", "street"}:
        return "GEOCODIFICADA_DIRECCION", "media"
    if place_rank >= 26 and addresstype not in {"town", "city", "village", "municipality"}:
        return "GEOCODIFICADA_DIRECCION", "media"
    if result_type == "administrative" or addresstype in {"town", "city", "village", "municipality", "city_district"}:
        return "CENTROIDE_LOCALIDAD", "media"
    return "GEOCODIFICADA_DIRECCION", "media"


def geocode_row(row, cache, delay, enable_geocode=True, mymaps_index=None):
    lat = to_float(row.get("LAT_DEC"))
    lon = to_float(row.get("LON_DEC"))
    localidad, provincia = infer_location(row.get("LOCALIDAD"), row.get("PROVINCIA"))
    direccion = build_address(row)

    if mymaps_index:
        mymaps_key = raw_match_key(
            row.get("DISTRIBUIDOR"),
            row.get("DOMICILIO"),
            row.get("LOCALIDAD"),
            row.get("PROVINCIA"),
        )
        mymaps_point = mymaps_index.get(mymaps_key)
        if mymaps_point:
            return {
                "lat": mymaps_point["lat"],
                "lon": mymaps_point["lon"],
                "estado": "EXACTA_MYMAPS",
                "confianza": "alta",
                "fuente": "Google My Maps",
                "consulta": raw_match_key(
                    row.get("DISTRIBUIDOR"),
                    row.get("DOMICILIO"),
                    row.get("LOCALIDAD"),
                    row.get("PROVINCIA"),
                ),
                "resultado": clean(mymaps_point.get("mymaps_id")),
            }

    if in_argentina(lat, lon):
        return {
            "lat": lat,
            "lon": lon,
            "estado": "EXACTA_PLANILLA",
            "confianza": "alta",
            "fuente": "Planilla Corteva LAT_DEC/LON_DEC",
            "consulta": "",
            "resultado": clean(row.get("COORDENADAS_DECIMALES")),
        }

    for field in ["GOOGLE_MAPS", "LINK_BUSQUEDA_GOOGLE_MAPS", "LINK_COMO_LLEGAR"]:
        coordinates = extract_coordinates_from_text(row.get(field))
        if coordinates:
            lat, lon = coordinates
            return {
                "lat": lat,
                "lon": lon,
                "estado": "EXACTA_GOOGLE_LINK",
                "confianza": "alta",
                "fuente": f"Excel Corteva {field}",
                "consulta": clean(row.get(field)),
                "resultado": f"{lat},{lon}",
            }

    if not enable_geocode:
        return {
            "lat": None,
            "lon": None,
            "estado": "PENDIENTE_GEOCODIFICAR",
            "confianza": "revisar",
            "fuente": "",
            "consulta": direccion,
            "resultado": "",
        }

    google_queries = []
    for field in ["GOOGLE_MAPS", "LINK_BUSQUEDA_GOOGLE_MAPS", "LINK_COMO_LLEGAR"]:
        query = extract_query_from_google_url(row.get(field))
        if query:
            google_queries.append(query)
    distribuidor = clean(row.get("DISTRIBUIDOR"))
    if distribuidor:
        google_queries.append(f"{distribuidor}, {direccion}")
    google_queries.append(direccion)

    seen_google_queries = set()
    for google_query in google_queries:
        query_key = clean(google_query).lower()
        if not query_key or query_key in seen_google_queries:
            continue
        seen_google_queries.add(query_key)
        google_place = google_places_search(google_query, localidad, provincia, cache, delay)
        if google_place:
            return google_place
        google_result = google_geocode(google_query, localidad, provincia, cache, delay)
        if google_result:
            return google_result

    address_queries = [direccion]
    address_norm = normalize(direccion)
    if "CIUDAD AUTONOMA BUENOS AIRES" in address_norm:
        domicilio = clean(row.get("DOMICILIO"))
        if domicilio:
            address_queries.append(f"{domicilio}, Ciudad Autonoma de Buenos Aires, Argentina")
    if "WENCESLAO ESCALANTE" in address_norm:
        address_queries.append(clean(direccion).replace("WENSESLAO", "WENCESLAO").replace("Wenseslao", "Wenceslao"))
    if "SALLIQUELO" in address_norm:
        address_queries.append(clean(direccion).replace("SALLIQUELLO", "SALLIQUELO").replace("Salliquello", "Salliquelo"))
    if normalize(localidad) == "CASA CENTRAL" and "ALFREDO GUZMAN" in address_norm:
        address_queries.extend(
            [
                "Av. Alfredo Guzman 10, Tucuman, Argentina",
                "Av. Alfredo Guzman 10, Alderetes, Tucuman, Argentina",
            ]
        )

    seen_queries = set()
    for address_query in address_queries:
        query_key = clean(address_query).lower()
        if query_key in seen_queries:
            continue
        seen_queries.add(query_key)
        address_results = nominatim_search(address_query, cache, delay)
        selected = select_nominatim_result(address_results, localidad, provincia, "address")
        if selected:
            estado, confianza = classify_nominatim_result(selected, "address")
            return {
                "lat": to_float(selected.get("lat")),
                "lon": to_float(selected.get("lon")),
                "estado": estado,
                "confianza": confianza,
                "fuente": "Nominatim / OpenStreetMap",
                "consulta": address_query,
                "resultado": clean(selected.get("display_name")),
            }

    localidad_query_parts = [localidad, provincia, "Argentina"]
    localidad_query = ", ".join([p for p in localidad_query_parts if p])
    locality_queries = [localidad_query]
    loc_norm = normalize(localidad)
    prov_norm = normalize(provincia)
    if loc_norm == "ALDERETES" and prov_norm == "TUCUMAN":
        locality_queries.append("Alderetes, Tucuman, Argentina")
    if loc_norm == "VILLA SANTA ROSA" and prov_norm == "CORDOBA":
        locality_queries.append("Villa Santa Rosa, Cordoba, Argentina")
    if loc_norm == "LA DORMIDA" and prov_norm == "CORDOBA":
        locality_queries.append("La Dormida, Cordoba, Argentina")
    if loc_norm == "RIO SEGUNDO" and prov_norm == "CORDOBA":
        locality_queries.append("Municipio de Rio Segundo, Cordoba, Argentina")
    if loc_norm == "SINSACATE" and prov_norm == "CORDOBA":
        locality_queries.append("Sinsacate, Totoral, Cordoba, Argentina")
    if loc_norm == "HUINCA RENANCO" and prov_norm == "CORDOBA":
        locality_queries.append("Huinca Renanco, General Roca, Cordoba, Argentina")
    if loc_norm == "CIUDAD AUTONOMA BUENOS AIRES":
        locality_queries.append("Buenos Aires, Ciudad Autonoma de Buenos Aires, Argentina")
    if loc_norm == "GENERAL PICO" and prov_norm == "LA PAMPA":
        locality_queries.append("General Pico, La Pampa, Argentina")
    if loc_norm == "JUAN JOSE CASTELLI" and prov_norm == "CHACO":
        locality_queries.append("Juan Jose Castelli, Chaco, Argentina")
    if loc_norm == "VILLA MARIA" and prov_norm == "CORDOBA":
        locality_queries.append("Villa Maria, Cordoba, Argentina")
    if loc_norm == "GENERAL BELGRANO" and prov_norm == "BUENOS AIRES":
        locality_queries.append("General Belgrano, Buenos Aires, Argentina")
    if loc_norm == "CORONEL SUAREZ" and prov_norm == "BUENOS AIRES":
        locality_queries.append("Coronel Suarez, Buenos Aires, Argentina")
    if loc_norm == "BOLIVAR" and prov_norm == "BUENOS AIRES":
        locality_queries.append("San Carlos de Bolivar, Buenos Aires, Argentina")
    if loc_norm in {"P R SAENZ PENA", "P R SAENZ PENA CHACO"}:
        locality_queries.append("Presidencia Roque Saenz Pena, Chaco, Argentina")

    seen_locality_queries = set()
    for locality_query in locality_queries:
        query_key = clean(locality_query).lower()
        if query_key in seen_locality_queries:
            continue
        seen_locality_queries.add(query_key)
        locality_results = nominatim_search(locality_query, cache, delay)
        selected = select_nominatim_result(locality_results, localidad, provincia, "locality")
        if selected:
            estado, confianza = classify_nominatim_result(selected, "locality")
            return {
                "lat": to_float(selected.get("lat")),
                "lon": to_float(selected.get("lon")),
                "estado": estado,
                "confianza": confianza,
                "fuente": "Nominatim / OpenStreetMap",
                "consulta": locality_query,
                "resultado": clean(selected.get("display_name")),
            }

    return {
        "lat": None,
        "lon": None,
        "estado": "REVISAR_SIN_COORDENADA",
        "confianza": "revisar",
        "fuente": "Sin resultado confiable",
        "consulta": direccion,
        "resultado": json.dumps(address_results, ensure_ascii=False)[:500],
    }


def build_payload(record, nombre, geocode):
    direccion = build_address(record)
    geojson = None
    if geocode["lat"] is not None and geocode["lon"] is not None:
        geojson = {
            "type": "Point",
            "coordinates": [round(geocode["lon"], 8), round(geocode["lat"], 8)],
        }
    return {
        "nombre": nombre,
        "direccion": direccion,
        "geojson": geojson,
        "idQuimica": "__CORTEVA_ID__",
    }


def write_csv(path, rows, headers):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--delay", type=float, default=1.1)
    parser.add_argument("--skip-geocode", action="store_true")
    parser.add_argument("--mymaps-points", default="")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    cache_path = output_dir / "geocode-cache.json"
    cache = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else {}

    records = load_records(input_path)
    if args.limit:
        records = records[: args.limit]
    names = make_unique_names(records)
    mymaps_index = load_mymaps_points(args.mymaps_points)
    output_rows = []
    payload = []

    for index, (record, nombre) in enumerate(zip(records, names), start=1):
        geocode = geocode_row(
            record,
            cache,
            args.delay,
            enable_geocode=not args.skip_geocode,
            mymaps_index=mymaps_index,
        )
        payload_item = build_payload(record, nombre, geocode)
        payload.append(payload_item)
        localidad, provincia = infer_location(record.get("LOCALIDAD"), record.get("PROVINCIA"))
        output_rows.append(
            {
                "fila_excel": record["_excel_row"],
                "nombre_chaman": nombre,
                "distribuidor_original": clean(record.get("DISTRIBUIDOR")),
                "localidad": localidad,
                "provincia": provincia,
                "direccion_chaman": payload_item["direccion"],
                "telefono": clean(record.get("TELEFONO")),
                "codigo_postal": clean(record.get("C.P.")),
                "lat": geocode["lat"],
                "lon": geocode["lon"],
                "estado_coordenada_chaman": geocode["estado"],
                "confianza": geocode["confianza"],
                "fuente_geocodificacion": geocode["fuente"],
                "consulta_geocodificacion": geocode["consulta"],
                "resultado_geocodificacion": geocode["resultado"],
                "google_maps": clean(record.get("GOOGLE_MAPS")),
                "link_busqueda_google_maps": clean(record.get("LINK_BUSQUEDA_GOOGLE_MAPS")),
                "link_como_llegar": clean(record.get("LINK_COMO_LLEGAR")),
                "observaciones_origen": clean(record.get("OBSERVACIONES")),
                "payload_json": json.dumps(payload_item, ensure_ascii=False),
            }
        )
        if index % 25 == 0:
            cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"Procesadas {index}/{len(records)} filas")

    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    headers = list(output_rows[0].keys()) if output_rows else []
    write_csv(output_dir / "corteva_distribuidores_geocoded.csv", output_rows, headers)
    (output_dir / "corteva_distribuidores_geocoded.json").write_text(
        json.dumps(output_rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "corteva_distribuidores_import_payload.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    summary = {
        "total": len(output_rows),
        "estados": Counter(row["estado_coordenada_chaman"] for row in output_rows),
        "confianza": Counter(row["confianza"] for row in output_rows),
        "sin_geojson": sum(1 for item in payload if not item["geojson"]),
        "nombres_unicos": len(set(normalize(row["nombre_chaman"]) for row in output_rows)),
    }
    summary = {
        **summary,
        "estados": dict(summary["estados"]),
        "confianza": dict(summary["confianza"]),
    }
    (output_dir / "corteva_distribuidores_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
